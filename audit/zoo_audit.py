#!/usr/bin/env python3
# =============================================================================
# glang — Topology Zoo GraphML audit generator
#
# Copyright 2026 Eric Parsonage
#
# This software is provided subject to the terms of the included licence.
#
# Author: Eric Parsonage
# =============================================================================
"""Topology Zoo GraphML audit.

Walks every *.graphml file under ~/InternetTopologyZoo/graphml, records
every node's and every edge's attributes, and produces:

  zoo-audit.xlsx  — one sheet per graph plus four cross-graph rollups
  report.md       — human-readable summary of overall data quality
  report.pdf      — typeset version of the same summary

Findings logged:

- nodes missing Latitude or Longitude (either absent, empty, "None",
  "NaN", "null", or outside the valid range)
- non-numeric coordinates where numeric was expected
- self-loops
- parallel edges (same unordered endpoint pair, any direction)
- duplicate node labels within a graph
- edges referring to undefined node ids
- isolated (unconnected) nodes
- nodes with unusual labels (empty, whitespace-only, or containing
  control characters)
- graphs flagged as multigraphs (edgedefault or presence of parallel
  edges)
- attribute-level oddities (mixed types, sentinel placeholders)

Usage:
  /home/eric/.venv/bin/python zoo_audit.py

Output lands in the same directory as this script.
"""
from __future__ import annotations
import os, sys, re, math, unicodedata, urllib.request, zipfile, io
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import numpy as np


# ---- config ---------------------------------------------------------

COPYRIGHT = "Copyright © 2026 Eric Parsonage"

ZOO_DIR = Path.home() / "InternetTopologyZoo" / "graphml"
OUT_DIR = Path(__file__).parent

# Geonames cities500 dataset: ~350k populated places, every entry has a
# population.  Used for two enrichment paths in the audit:
#   * REVERSE: a node with lat/lon → its nearest cities500 record's
#              country/name/population (no country attr needed).
#   * FORWARD: a node lacking lat/lon but with a label and a country
#              scope → the matching cities500 record by name within
#              that country.
# Cached under ~/.cache/glang/geonames/ so re-runs of the audit don't
# re-download.
GEONAMES_URL    = "https://download.geonames.org/export/dump/cities500.zip"
GEONAMES_CACHE  = Path.home() / ".cache" / "glang" / "geonames"


# ---- geonames index ------------------------------------------------

@dataclass
class GeoRecord:
    geonameid: int
    name: str
    asciiname: str
    lat: float
    lon: float
    country: str   # ISO-2 e.g. "AU"
    population: int


class GeonamesIndex:
    """In-memory cities500 index supporting reverse (lat/lon → record)
    and forward (country + name → record) lookup.  Numpy-backed so a
    full 9.7k-node sweep finishes in seconds.
    """

    @classmethod
    def load(cls):
        path = GEONAMES_CACHE / "cities500.txt"
        if not path.exists():
            GEONAMES_CACHE.mkdir(parents=True, exist_ok=True)
            print(f"  fetching {GEONAMES_URL} ...", flush=True)
            with urllib.request.urlopen(GEONAMES_URL, timeout=60) as resp:
                data = resp.read()
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                zf.extract("cities500.txt", path=GEONAMES_CACHE)
        return cls(path)

    def __init__(self, path: Path):
        records: list[GeoRecord] = []
        with path.open(encoding="utf-8") as f:
            for line in f:
                cols = line.rstrip("\n").split("\t")
                if len(cols) < 19:
                    continue
                try:
                    records.append(GeoRecord(
                        geonameid=int(cols[0]),
                        name=cols[1],
                        asciiname=cols[2],
                        lat=float(cols[4]),
                        lon=float(cols[5]),
                        country=cols[8],
                        population=int(cols[14] or 0),
                    ))
                except (ValueError, IndexError):
                    continue
        self.records = records
        # Numpy arrays for fast reverse-lookup vectorised haversine.
        # Pre-radian conversion avoids per-query overhead.
        self.lat_rad = np.array([math.radians(r.lat) for r in records])
        self.lon_rad = np.array([math.radians(r.lon) for r in records])
        # Forward index: country + lower-cased name (and asciiname) → record.
        # If multiple cities share a name we keep the most-populous one.
        self.by_country_name: dict[tuple[str, str], GeoRecord] = {}
        for r in records:
            for nm in {r.name.lower(), r.asciiname.lower()}:
                if not nm:
                    continue
                key = (r.country, nm)
                cur = self.by_country_name.get(key)
                if cur is None or r.population > cur.population:
                    self.by_country_name[key] = r

    def reverse(self, lat: float, lon: float) -> GeoRecord:
        """Nearest cities500 record by haversine distance."""
        lat_r = math.radians(lat)
        lon_r = math.radians(lon)
        dlat = self.lat_rad - lat_r
        dlon = self.lon_rad - lon_r
        a = np.sin(dlat * 0.5) ** 2 + np.cos(self.lat_rad) * math.cos(lat_r) * np.sin(dlon * 0.5) ** 2
        # Argmin of sqrt-haversine is argmin of haversine; skip sqrt.
        i = int(np.argmin(a))
        return self.records[i]

    def forward(self, name: str, country: str) -> GeoRecord | None:
        """Best name match within the given ISO-2 country, or None.
        Tries the literal label first, then a unicode-stripped form."""
        if not name or not country:
            return None
        cc = country.upper()
        nm = name.strip().lower()
        rec = self.by_country_name.get((cc, nm))
        if rec is not None:
            return rec
        # Try unicode-NFD-stripped form (Müller → Muller).
        ascii_nm = "".join(
            c for c in unicodedata.normalize("NFD", nm)
            if unicodedata.category(c) != "Mn"
        )
        return self.by_country_name.get((cc, ascii_nm))


# Country-name → ISO-2 for graph-level GeoLocation values like
# "Australia".  Topology Zoo uses a small set; we cover the ones that
# appear in single-country-scoped graphs.  Unknown values just yield
# None and the per-node Country attribute remains the only source.
_COUNTRY_ISO2 = {
    "argentina": "AR", "australia": "AU", "austria": "AT", "belgium": "BE",
    "brazil": "BR", "bulgaria": "BG", "canada": "CA", "chile": "CL",
    "china": "CN", "colombia": "CO", "croatia": "HR", "cyprus": "CY",
    "czech republic": "CZ", "denmark": "DK", "egypt": "EG", "estonia": "EE",
    "finland": "FI", "france": "FR", "georgia": "GE", "germany": "DE",
    "ghana": "GH", "greece": "GR", "hong kong": "HK", "hungary": "HU",
    "iceland": "IS", "india": "IN", "indonesia": "ID", "iran": "IR",
    "ireland": "IE", "israel": "IL", "italy": "IT", "japan": "JP",
    "kenya": "KE", "latvia": "LV", "lithuania": "LT", "luxembourg": "LU",
    "macedonia": "MK", "malaysia": "MY", "malta": "MT", "mexico": "MX",
    "moldova": "MD", "mongolia": "MN", "morocco": "MA", "netherlands": "NL",
    "new zealand": "NZ", "nigeria": "NG", "norway": "NO", "pakistan": "PK",
    "peru": "PE", "philippines": "PH", "poland": "PL", "portugal": "PT",
    "romania": "RO", "russia": "RU", "saudi arabia": "SA", "serbia": "RS",
    "singapore": "SG", "slovakia": "SK", "slovenia": "SI", "south africa": "ZA",
    "south korea": "KR", "spain": "ES", "sri lanka": "LK", "sweden": "SE",
    "switzerland": "CH", "taiwan": "TW", "thailand": "TH", "tunisia": "TN",
    "turkey": "TR", "uganda": "UG", "ukraine": "UA", "uk": "GB",
    "united kingdom": "GB", "united states": "US", "uruguay": "UY",
    "usa": "US", "venezuela": "VE", "vietnam": "VN", "zambia": "ZM",
}


def country_to_iso2(name: str | None) -> str | None:
    if not name:
        return None
    return _COUNTRY_ISO2.get(name.strip().lower())


XLSX    = OUT_DIR / "zoo-audit.xlsx"
REPORT  = OUT_DIR / "report.md"
PDF     = OUT_DIR / "report.pdf"

NS = "{http://graphml.graphdrawing.org/xmlns}"

MISSING_SENTINELS = {"", "none", "nan", "null", "undef", "undefined", "-",
                     "na", "n/a", "0.0.0.0"}


# ---- parsing helpers ------------------------------------------------

def tag(elem: ET.Element) -> str:
    return elem.tag.replace(NS, "")


def parse_graphml(path: Path) -> dict:
    """Return a dict describing the file — see call sites for shape."""
    tree = ET.parse(path)
    root = tree.getroot()
    if tag(root) != "graphml":
        raise ValueError(f"{path}: root not <graphml>")

    keys = {}   # key-id -> (attr-name, target)
    for k in root.findall(NS + "key"):
        kid  = k.get("id")
        name = k.get("attr.name") or kid
        trg  = k.get("for") or "all"
        typ  = k.get("attr.type") or "string"
        if kid:
            keys[kid] = {"name": name, "for": trg, "type": typ}

    g = root.find(NS + "graph")
    if g is None:
        raise ValueError(f"{path}: no <graph>")

    edgedefault = g.get("edgedefault") or "directed"
    parsenodeids = g.get("parse.nodeids") or ""

    def data_of(elem, target):
        out = {}
        for d in elem.findall(NS + "data"):
            kid = d.get("key")
            if not kid:
                continue
            key = keys.get(kid)
            if not key:
                continue
            if target and key["for"] not in (target, "all"):
                continue
            out[key["name"]] = (d.text or "").strip()
        return out

    gdata  = data_of(g, "graph")
    nodes  = []
    for n in g.findall(NS + "node"):
        nodes.append({"id": n.get("id"), "attrs": data_of(n, "node")})
    edges  = []
    for e in g.findall(NS + "edge"):
        edges.append({
            "source": e.get("source"),
            "target": e.get("target"),
            "id":     e.get("id"),
            "attrs":  data_of(e, "edge"),
        })

    return {
        "path":         path,
        "name":         path.stem,
        "keys":         keys,
        "graph_data":   gdata,
        "edgedefault":  edgedefault,
        "parse_nodeids": parsenodeids,
        "nodes":        nodes,
        "edges":        edges,
    }


# ---- normalisation / predicate helpers ------------------------------

def is_missing(v) -> bool:
    if v is None:
        return True
    if not isinstance(v, str):
        return False
    s = v.strip().lower()
    return s in MISSING_SENTINELS


def coerce_float(v):
    if is_missing(v):
        return None
    try:
        x = float(v)
        if math.isnan(x) or math.isinf(x):
            return None
        return x
    except (TypeError, ValueError):
        return None


def has_control_chars(s: str) -> bool:
    if not s:
        return False
    for ch in s:
        if unicodedata.category(ch).startswith("C") and ch not in ("\t", "\n"):
            return True
    return False


# ---- per-graph analysis --------------------------------------------

def analyse(parsed: dict, geo: GeonamesIndex | None = None):
    """Return (per-graph summary, issues list, missing-latlon list,
    real-node enrichment list).
    When `geo` is supplied, every real internal node gets a geonames
    lookup attempt: reverse from lat/lon when present, otherwise
    forward by name+country.  The matching record's country, place
    name and population are attached to the per-node enrichment row.
    """
    name  = parsed["name"]
    nodes = parsed["nodes"]
    edges = parsed["edges"]
    issues = []
    enrichment_rows = []   # (graph, node_id, label, src_lat, src_lon,
                           #  matched_country, matched_name, population, method)
    real_resolved   = 0
    real_total_pop  = 0

    # --- nodes ---
    node_ids = {n["id"] for n in nodes}
    node_label_counts = Counter()
    missing_latlon = []  # (node_id, label, which)
    bad_coords = []      # (node_id, label, kind, raw)
    label_empty = 0
    label_whitespace = 0
    label_control = 0

    # Topology Zoo's graph-level country context lives in two attrs:
    #   GeoExtent   — categorical: "Country", "Region", "Continent",
    #                 "Global".  Only "Country" narrows the search to a
    #                 single geonames-per-country file usefully.
    #   GeoLocation — the actual place name (e.g. "Australia").
    # If GeoExtent is "Country" with a non-empty GeoLocation we treat
    # the graph as single-country-scoped; otherwise per-node Country
    # attrs are required for geonames forward-lookup to resolve.
    graph_data        = parsed.get("graph_data", {}) or {}
    geo_extent        = (graph_data.get("GeoExtent") or "").strip()
    geo_location      = (graph_data.get("GeoLocation") or "").strip()
    graph_is_single_country = (
        geo_extent == "Country"
        and geo_location
        and not is_missing(geo_location))

    # Real-node accounting: glang's topology-zoo plugin classifies any
    # node with Internal==0 as Rj45 (external/peering stub) and any
    # node with hyperedge==1 as a switch.  Everything else is a "real"
    # node that becomes a CORE router and ought to have geocoordinates
    # for honest map placement.  We track real-node lat/lon gaps
    # separately because a missing coordinate on an Rj45 stub is fine
    # (geonames spring-places it from neighbouring anchors); a missing
    # coordinate on a real router is more informative about the source
    # data quality.
    rj45_nodes              = 0
    rj45_node_ids           = set()
    switch_nodes            = 0
    real_nodes              = 0
    real_missing_latlon     = 0
    real_missing_label      = 0
    real_missing_country    = 0
    # Per-node: had no lat/lon, AND would also have no name+country
    # fallback for forward-lookup -- so geonames couldn't reach this
    # node by either route.  Aggregated up at the end into the gate.
    real_geonames_unreachable = 0

    for n in nodes:
        a = n["attrs"]
        nid = n["id"]
        label = a.get("label", "")
        node_label_counts[label.strip()] += 1

        if not label:
            label_empty += 1
            issues.append({"category": "node-label-missing", "severity": "warn",
                           "detail": f"node id={nid} has no label attribute"})
        elif label.strip() == "":
            label_whitespace += 1
            issues.append({"category": "node-label-whitespace", "severity": "warn",
                           "detail": f"node id={nid} label is whitespace-only: {label!r}"})
        elif has_control_chars(label):
            label_control += 1
            issues.append({"category": "node-label-control", "severity": "warn",
                           "detail": f"node id={nid} label contains control characters: {label!r}"})

        lat = a.get("Latitude")
        lon = a.get("Longitude")
        fl_lat = coerce_float(lat)
        fl_lon = coerce_float(lon)
        is_unlocated = fl_lat is None or fl_lon is None

        if fl_lat is None and fl_lon is None:
            missing_latlon.append((nid, label, "both"))
        elif fl_lat is None:
            missing_latlon.append((nid, label, "lat"))
            if lat is not None and not is_missing(lat):
                bad_coords.append((nid, label, "Latitude", lat))
        elif fl_lon is None:
            missing_latlon.append((nid, label, "lon"))
            if lon is not None and not is_missing(lon):
                bad_coords.append((nid, label, "Longitude", lon))

        # Classify per glang's default topology-zoo shape rules.
        internal_raw  = a.get("Internal")
        hyperedge_raw = a.get("hyperedge")
        country_raw   = a.get("Country")
        if str(internal_raw) == "0":
            rj45_nodes += 1
            rj45_node_ids.add(nid)
        elif str(hyperedge_raw) == "1":
            switch_nodes += 1
        else:
            real_nodes += 1
            if is_unlocated:
                real_missing_latlon += 1
                issues.append({"category": "real-node-missing-latlon", "severity": "warn",
                               "detail": f"real node id={nid} ({label!r}) has no Latitude/Longitude"})
            # Per-node geonames reachability:
            #
            # Geonames offers two lookup paths.  A node is reachable if
            # at least one applies.
            #
            #   reverse: lat/lon → nearest populated place.  No country
            #            needed (the coordinates disambiguate globally).
            #
            #   forward: place name → matching geonames record.  Needs
            #            a country scope to point at the right per-country
            #            data file -- either the node's own `Country`
            #            attribute or a graph-level single-country
            #            context (GeoExtent == "Country" + GeoLocation).
            #
            # A node is "unreachable" only when BOTH paths fail: it has
            # no lat/lon AND has no name+country fallback.
            label_ok = bool(label and label.strip()
                            and not has_control_chars(label))
            country_ok = (
                graph_is_single_country
                or (country_raw is not None and not is_missing(country_raw)))
            if not label_ok:
                real_missing_label += 1
            if country_raw is None or is_missing(country_raw):
                real_missing_country += 1
            forward_ok = label_ok and country_ok
            reverse_ok = not is_unlocated
            if not (forward_ok or reverse_ok):
                real_geonames_unreachable += 1

            # Live geonames enrichment: reverse from lat/lon when
            # present (no country needed), else forward from
            # name+country.  The matching record's country/name/
            # population is captured per node and aggregated.
            if geo is not None:
                rec = None
                method = "unreachable"
                if reverse_ok:
                    rec = geo.reverse(fl_lat, fl_lon)
                    method = "reverse"
                elif forward_ok:
                    cc = country_to_iso2(country_raw) \
                        or (country_to_iso2(geo_location)
                            if graph_is_single_country else None)
                    if cc:
                        rec = geo.forward(label, cc)
                        if rec is not None:
                            method = "forward"
                        else:
                            method = "forward-miss"
                if rec is not None:
                    real_resolved += 1
                    real_total_pop += rec.population
                    enrichment_rows.append((
                        name, nid, label,
                        fl_lat, fl_lon,
                        rec.country, rec.name, rec.population, method))
                else:
                    enrichment_rows.append((
                        name, nid, label,
                        fl_lat, fl_lon,
                        "", "", 0, method))

        if fl_lat is not None and not (-90.0 <= fl_lat <= 90.0):
            bad_coords.append((nid, label, "Latitude", lat))
            issues.append({"category": "node-lat-out-of-range", "severity": "error",
                           "detail": f"node id={nid} ({label!r}) lat={lat} outside [-90, 90]"})
        if fl_lon is not None and not (-180.0 <= fl_lon <= 180.0):
            bad_coords.append((nid, label, "Longitude", lon))
            issues.append({"category": "node-lon-out-of-range", "severity": "error",
                           "detail": f"node id={nid} ({label!r}) lon={lon} outside [-180, 180]"})

    for lbl, count in node_label_counts.items():
        if count > 1 and lbl:  # whitespace already flagged
            issues.append({"category": "node-label-duplicate", "severity": "warn",
                           "detail": f"label {lbl!r} used by {count} nodes"})

    for nid, label, which in missing_latlon:
        issues.append({"category": f"node-missing-{which}", "severity": "info",
                       "detail": f"node id={nid} ({label!r}) missing {which}"})

    # --- edges ---
    undirected_pair_counts = Counter()
    self_loops = 0
    dangling_src = 0
    dangling_dst = 0
    external_edges = 0     # edges with at least one endpoint classified as Rj45
    for e in edges:
        s, t = e["source"], e["target"]
        if s not in node_ids:
            dangling_src += 1
            issues.append({"category": "edge-dangling-source", "severity": "error",
                           "detail": f"edge id={e['id']} source={s} not in node set"})
        if t not in node_ids:
            dangling_dst += 1
            issues.append({"category": "edge-dangling-target", "severity": "error",
                           "detail": f"edge id={e['id']} target={t} not in node set"})
        if s == t:
            self_loops += 1
            issues.append({"category": "edge-self-loop", "severity": "warn",
                           "detail": f"edge id={e['id']} source=target={s}"})
        # An edge is external (a connection to another network) when
        # at least one endpoint was classified as Rj45 above.  This is
        # the count of cross-domain peering / transit links the zoo
        # operator drew into the diagram.
        if s in rj45_node_ids or t in rj45_node_ids:
            external_edges += 1
        unordered = tuple(sorted((s, t))) if s and t else (s, t)
        undirected_pair_counts[unordered] += 1

    parallel_edges = sum(c - 1 for c in undirected_pair_counts.values() if c > 1)
    if parallel_edges:
        for pair, c in undirected_pair_counts.items():
            if c > 1:
                issues.append({"category": "edge-parallel", "severity": "info",
                               "detail": f"pair {pair} has {c} edges"})

    # --- connectivity ---
    adj = defaultdict(set)
    for e in edges:
        if e["source"] in node_ids and e["target"] in node_ids:
            adj[e["source"]].add(e["target"])
            adj[e["target"]].add(e["source"])
    isolated_nodes = [n["id"] for n in nodes if n["id"] not in adj]
    for nid in isolated_nodes:
        issues.append({"category": "node-isolated", "severity": "warn",
                       "detail": f"node id={nid} has no edges"})

    # components via BFS
    visited = set()
    components = 0
    for n in nodes:
        nid = n["id"]
        if nid in visited:
            continue
        components += 1
        frontier = [nid]
        while frontier:
            cur = frontier.pop()
            if cur in visited:
                continue
            visited.add(cur)
            frontier.extend(adj.get(cur, ()))
    if components > 1:
        issues.append({"category": "graph-disconnected", "severity": "info",
                       "detail": f"{components} connected components"})

    # multigraph? edgedefault or any parallel
    declared_multi = parsed["edgedefault"] == "undirected" and False  # edgedefault alone isn't multi
    is_multi = parallel_edges > 0

    # Could geonames potentially resolve every internal node and
    # contribute population data?  Two gating conditions:
    #
    #   - Every real node has a non-empty label (geonames searches by
    #     place-name string).
    #   - Every real node is scoped to a country, either via a
    #     per-node `Country` attribute or via a graph-level
    #     `GeoExtent`/`Country` attribute that covers the whole
    #     topology.
    #
    # If both conditions hold we count the topology as "geonames
    # eligible".  Whether geonames actually finds each label in the
    # corresponding country file is a separate, more expensive check
    # (running the lookup against the live data files) -- this gate
    # is the cheap upper bound.
    # A graph is "geonames-eligible" when every real internal node is
    # reachable by either of geonames' two lookup paths -- reverse
    # (from lat/lon) or forward (from place name + country scope).
    # The per-node check ran inside the loop above and tallied
    # `real_geonames_unreachable`; the graph-level gate is just
    # "no real nodes were unreachable".
    geonames_eligible = (real_nodes > 0
                         and real_geonames_unreachable == 0)

    summary = {
        "graph":                name,
        "nodes":                len(nodes),
        "real_nodes":           real_nodes,
        "rj45_nodes":           rj45_nodes,
        "switch_nodes":         switch_nodes,
        "edges":                len(edges),
        "external_edges":       external_edges,
        "edgedefault":          parsed["edgedefault"],
        "multigraph?":          "yes" if is_multi else "no",
        "parallel_edges":       parallel_edges,
        "self_loops":           self_loops,
        "missing_lat_or_lon":           len(missing_latlon),
        "real_missing_lat_or_lon":      real_missing_latlon,
        "real_missing_label":           real_missing_label,
        "real_missing_country":         real_missing_country,
        "real_geonames_unreachable":    real_geonames_unreachable,
        "geonames_eligible":            "yes" if geonames_eligible else "no",
        "real_nodes_resolved":          real_resolved,
        "real_total_population":        real_total_pop,
        "bad_coords":           len(bad_coords),
        "isolated_nodes":       len(isolated_nodes),
        "components":           components,
        "duplicate_labels":     sum(1 for c in node_label_counts.values() if c > 1),
        "whitespace_labels":    label_whitespace,
        "empty_labels":         label_empty,
        "control_char_labels":  label_control,
        "dangling_endpoints":   dangling_src + dangling_dst,
        "issues":               len(issues),
    }
    return summary, issues, missing_latlon, enrichment_rows


# ---- xlsx output ----------------------------------------------------

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
ISSUE_FILL  = PatternFill("solid", fgColor="FFE5B4")


def style_header(ws, row=1):
    for c in ws[row]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL


def autosize(ws):
    for col in ws.columns:
        if not col:
            continue
        try:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        except ValueError:
            length = 10
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 40)


def write_per_graph_sheet(wb, parsed):
    # Sheet name must be <=31 chars and contain no : \ / ? * [ ]
    name = re.sub(r"[:\\/?*\[\]]", "_", parsed["name"])[:31]
    if name in wb.sheetnames:
        name = (name[:28] + "_2")[:31]
    ws = wb.create_sheet(name)

    ws["A1"] = f"Graph: {parsed['name']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = COPYRIGHT
    ws["A2"].font = Font(italic=True, color="666666")

    row = 4
    ws.cell(row, 1, "Graph-level data").font = HEADER_FONT
    row += 1
    ws.cell(row, 1, "edgedefault"); ws.cell(row, 2, parsed["edgedefault"]); row += 1
    for k, v in sorted(parsed["graph_data"].items()):
        ws.cell(row, 1, k); ws.cell(row, 2, v); row += 1

    # --- nodes ---
    row += 1
    ws.cell(row, 1, f"Nodes ({len(parsed['nodes'])})").font = HEADER_FONT
    row += 1

    node_attr_keys = sorted({k for n in parsed["nodes"] for k in n["attrs"].keys()},
                            key=lambda x: (x != "label", x != "Latitude", x != "Longitude", x))
    headers = ["id", *node_attr_keys, "issues"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row, i, h).font = HEADER_FONT
        ws.cell(row, i).fill = HEADER_FILL
    row += 1
    for n in parsed["nodes"]:
        nid = n["id"]
        a = n["attrs"]
        node_issues = []
        lat_raw = a.get("Latitude")
        lon_raw = a.get("Longitude")
        fl_lat = coerce_float(lat_raw)
        fl_lon = coerce_float(lon_raw)
        if fl_lat is None:
            node_issues.append("missing lat" if is_missing(lat_raw) else f"bad lat {lat_raw!r}")
        if fl_lon is None:
            node_issues.append("missing lon" if is_missing(lon_raw) else f"bad lon {lon_raw!r}")
        if fl_lat is not None and not (-90 <= fl_lat <= 90):
            node_issues.append(f"lat out of range {lat_raw}")
        if fl_lon is not None and not (-180 <= fl_lon <= 180):
            node_issues.append(f"lon out of range {lon_raw}")
        if not a.get("label"):
            node_issues.append("no label")
        elif has_control_chars(a["label"]):
            node_issues.append("control chars in label")

        ws.cell(row, 1, nid)
        for i, k in enumerate(node_attr_keys, start=2):
            ws.cell(row, i, a.get(k, ""))
        ws.cell(row, len(headers), "; ".join(node_issues))
        if node_issues:
            for c in ws[row]:
                c.fill = ISSUE_FILL
        row += 1

    # --- edges ---
    row += 1
    ws.cell(row, 1, f"Edges ({len(parsed['edges'])})").font = HEADER_FONT
    row += 1

    edge_attr_keys = sorted({k for e in parsed["edges"] for k in e["attrs"].keys()})
    headers = ["source", "target", "id", *edge_attr_keys, "issues"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row, i, h).font = HEADER_FONT
        ws.cell(row, i).fill = HEADER_FILL
    row += 1

    nids = {n["id"] for n in parsed["nodes"]}
    pair_counts = Counter()
    for e in parsed["edges"]:
        s, t = e["source"], e["target"]
        pair_counts[tuple(sorted((s, t))) if s and t else (s, t)] += 1

    for e in parsed["edges"]:
        edge_issues = []
        s, t = e["source"], e["target"]
        if s not in nids:
            edge_issues.append(f"dangling source {s}")
        if t not in nids:
            edge_issues.append(f"dangling target {t}")
        if s == t:
            edge_issues.append("self-loop")
        if pair_counts[tuple(sorted((s, t))) if s and t else (s, t)] > 1:
            edge_issues.append("parallel")

        ws.cell(row, 1, s)
        ws.cell(row, 2, t)
        ws.cell(row, 3, e["id"])
        for i, k in enumerate(edge_attr_keys, start=4):
            ws.cell(row, i, e["attrs"].get(k, ""))
        ws.cell(row, len(headers), "; ".join(edge_issues))
        if edge_issues:
            for c in ws[row]:
                c.fill = ISSUE_FILL
        row += 1

    autosize(ws)


def _stamp(ws, title):
    ws.cell(1, 1, title).font = Font(bold=True, size=14)
    ws.cell(2, 1, COPYRIGHT).font = Font(italic=True, color="666666")


def write_workbook(entries, all_missing, all_issues, all_enrichment,
                   global_summary):
    wb = Workbook()
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    _stamp(ws, "Topology Zoo — Summary")
    header_row = 4
    sum_cols = [
        "graph", "nodes", "real_nodes", "rj45_nodes", "switch_nodes",
        "edges", "external_edges", "edgedefault", "multigraph?",
        "parallel_edges", "self_loops", "missing_lat_or_lon",
        "real_missing_lat_or_lon", "real_missing_label",
        "real_missing_country", "real_geonames_unreachable",
        "geonames_eligible",
        "real_nodes_resolved", "real_total_population",
        "bad_coords", "isolated_nodes", "components",
        "duplicate_labels", "whitespace_labels", "empty_labels",
        "control_char_labels", "dangling_endpoints", "issues",
    ]
    for i, h in enumerate(sum_cols, start=1):
        ws.cell(header_row, i, h)
    style_header(ws, row=header_row)
    for r, e in enumerate(entries, start=header_row + 1):
        for i, k in enumerate(sum_cols, start=1):
            ws.cell(r, i, e.get(k, ""))
    # Totals row
    total_row = header_row + 1 + len(entries)
    ws.cell(total_row, 1, "TOTALS").font = HEADER_FONT
    for i, k in enumerate(sum_cols, start=1):
        if k in ("graph", "edgedefault", "multigraph?", "geonames_eligible"):
            continue
        vals = [e.get(k, 0) for e in entries if isinstance(e.get(k), (int, float))]
        ws.cell(total_row, i, sum(vals))
        ws.cell(total_row, i).font = HEADER_FONT
    autosize(ws)

    # Missing locations sheet
    ws2 = wb.create_sheet("Missing Locations")
    _stamp(ws2, "Topology Zoo — Missing Locations")
    ws2.cell(4, 1, "graph"); ws2.cell(4, 2, "node_id"); ws2.cell(4, 3, "label"); ws2.cell(4, 4, "missing")
    style_header(ws2, row=4)
    r = 5
    for graph, rows in all_missing.items():
        for (nid, label, which) in rows:
            ws2.cell(r, 1, graph)
            ws2.cell(r, 2, nid)
            ws2.cell(r, 3, label)
            ws2.cell(r, 4, which)
            r += 1
    autosize(ws2)

    # Issues sheet
    ws3 = wb.create_sheet("Issues")
    _stamp(ws3, "Topology Zoo — Issues")
    ws3.cell(4, 1, "graph"); ws3.cell(4, 2, "severity"); ws3.cell(4, 3, "category"); ws3.cell(4, 4, "detail")
    style_header(ws3, row=4)
    r = 5
    for graph, issues in all_issues.items():
        for iss in issues:
            ws3.cell(r, 1, graph)
            ws3.cell(r, 2, iss["severity"])
            ws3.cell(r, 3, iss["category"])
            ws3.cell(r, 4, iss["detail"])
            r += 1
    autosize(ws3)

    # Real-node populations sheet -- one row per real internal node,
    # showing the matched geonames record and population.
    ws_pop = wb.create_sheet("Real-node Populations")
    _stamp(ws_pop, "Topology Zoo — Real-node Populations")
    pop_headers = ["graph", "node_id", "label", "src_lat", "src_lon",
                   "matched_country", "matched_name", "population", "method"]
    for i, h in enumerate(pop_headers, start=1):
        ws_pop.cell(4, i, h)
    style_header(ws_pop, row=4)
    r = 5
    for graph, rows in all_enrichment.items():
        for row in rows:
            for i, v in enumerate(row, start=1):
                ws_pop.cell(r, i, v)
            r += 1
    autosize(ws_pop)

    # Partial Resolution sheet -- a to-do list of graphs that resolved
    # most of their real internal nodes but missed a handful.  Each
    # graph block opens with a header row showing the completion stats
    # and is followed by one row per unresolved node so a human can
    # work through them.  Sorted by completion percentage descending so
    # the easiest fixes come first.
    ws_part = wb.create_sheet("Partial Resolution")
    _stamp(ws_part, "Topology Zoo — Partially Resolved Graphs")
    part_headers = ["graph", "node_id", "label", "src_lat", "src_lon",
                    "method"]
    for i, h in enumerate(part_headers, start=1):
        ws_part.cell(4, i, h)
    style_header(ws_part, row=4)

    # Build a per-graph view from entries (so we have the stats handy).
    by_graph = {e["graph"]: e for e in entries}
    partials = []
    for g, e in by_graph.items():
        n = e.get("real_nodes", 0)
        k = e.get("real_nodes_resolved", 0)
        if isinstance(n, int) and isinstance(k, int) and 0 < k < n:
            partials.append((g, n, k))
    partials.sort(key=lambda t: (-t[2] / t[1], t[0]))

    r = 5
    for g, n, k in partials:
        pct = k * 100.0 / n
        hdr_cell = ws_part.cell(r, 1, f"{g}  —  {k}/{n} resolved ({pct:.1f}%)")
        hdr_cell.font = HEADER_FONT
        hdr_cell.fill = HEADER_FILL
        for col in range(2, len(part_headers) + 1):
            ws_part.cell(r, col).fill = HEADER_FILL
        r += 1
        for row in all_enrichment.get(g, []):
            (_, nid, label, src_lat, src_lon,
             matched_country, matched_name, _pop, method) = row
            if matched_name:   # resolved -- skip
                continue
            ws_part.cell(r, 1, g)
            ws_part.cell(r, 2, nid)
            ws_part.cell(r, 3, label)
            ws_part.cell(r, 4, src_lat)
            ws_part.cell(r, 5, src_lon)
            ws_part.cell(r, 6, method)
            for c in ws_part[r]:
                c.fill = ISSUE_FILL
            r += 1
    autosize(ws_part)

    # Multigraphs sheet
    ws4 = wb.create_sheet("Multigraphs")
    _stamp(ws4, "Topology Zoo — Multigraphs")
    ws4.cell(4, 1, "graph"); ws4.cell(4, 2, "nodes"); ws4.cell(4, 3, "edges")
    ws4.cell(4, 4, "parallel_edges"); ws4.cell(4, 5, "self_loops")
    style_header(ws4, row=4)
    r = 5
    for e in entries:
        if e["multigraph?"] == "yes":
            ws4.cell(r, 1, e["graph"])
            ws4.cell(r, 2, e["nodes"])
            ws4.cell(r, 3, e["edges"])
            ws4.cell(r, 4, e["parallel_edges"])
            ws4.cell(r, 5, e["self_loops"])
            r += 1
    autosize(ws4)

    return wb


# ---- markdown summary ------------------------------------------------

def write_report(entries, all_issues, all_enrichment, global_stats):
    total_graphs  = len(entries)
    total_nodes   = sum(e["nodes"] for e in entries)
    total_edges   = sum(e["edges"] for e in entries)
    total_missing = sum(e["missing_lat_or_lon"] for e in entries)
    total_self    = sum(e["self_loops"] for e in entries)
    total_para    = sum(e["parallel_edges"] for e in entries)
    total_iso     = sum(e["isolated_nodes"] for e in entries)
    total_bad     = sum(e["bad_coords"] for e in entries)
    total_dup     = sum(e["duplicate_labels"] for e in entries)
    total_dang    = sum(e["dangling_endpoints"] for e in entries)
    total_empty   = sum(e["empty_labels"] for e in entries)
    total_ws      = sum(e["whitespace_labels"] for e in entries)
    total_ctrl    = sum(e["control_char_labels"] for e in entries)
    multigraphs   = [e for e in entries if e["multigraph?"] == "yes"]

    completely_located  = sum(1 for e in entries if e["missing_lat_or_lon"] == 0)
    completely_unlocated = sum(1 for e in entries if e["missing_lat_or_lon"] == e["nodes"] and e["nodes"] > 0)
    partially_located    = total_graphs - completely_located - completely_unlocated

    # Issue category rollup
    category_counts = Counter()
    severity_counts = Counter()
    for issues in all_issues.values():
        for i in issues:
            category_counts[i["category"]] += 1
            severity_counts[i["severity"]] += 1

    # Graph sorted lists
    most_nodes = sorted(entries, key=lambda e: e["nodes"], reverse=True)[:10]
    most_missing = sorted(entries, key=lambda e: e["missing_lat_or_lon"], reverse=True)[:10]
    most_multi = sorted(multigraphs, key=lambda e: e["parallel_edges"], reverse=True)[:10]

    lines = []
    A = lines.append
    A("# Topology Zoo data-quality audit\n")
    A(f"*{COPYRIGHT}*\n")
    A(f"Source: `{ZOO_DIR}`  \n")
    A(f"Graphs parsed: **{total_graphs}**  \n")
    A(f"Total nodes: **{total_nodes:,}**  \n")
    A(f"Total edges: **{total_edges:,}**  \n")
    A("")

    A("## Location coverage\n")
    A(f"- **{completely_located}** graphs have every node located (lat + lon).")
    A(f"- **{partially_located}** graphs have a mix of located and unlocated nodes.")
    A(f"- **{completely_unlocated}** graphs have zero located nodes.")
    A(f"- Total nodes missing at least one of lat/lon: **{total_missing:,}**")
    A(f"  ({total_missing * 100.0 / max(total_nodes, 1):.1f}% of all nodes).")
    A(f"- Nodes whose lat/lon value was *provided but invalid* "
      f"(empty string, 'None', 'NaN', out-of-range, non-numeric): **{total_bad}**.")
    A("")

    # Real-vs-Rj45/switch breakdown.  Real nodes are CORE routers in
    # the exported XML; Rj45 nodes represent peering links to other
    # networks (per glang's default topology-zoo shape rules); switch
    # nodes represent shared-medium fabrics.
    total_real     = sum(e["real_nodes"]   for e in entries)
    total_rj45     = sum(e["rj45_nodes"]   for e in entries)
    total_switch   = sum(e["switch_nodes"] for e in entries)
    total_external = sum(e["external_edges"] for e in entries)
    total_real_unloc = sum(e["real_missing_lat_or_lon"] for e in entries)
    real_unloc_graphs = sum(1 for e in entries if e["real_missing_lat_or_lon"] > 0)
    geon_eligible_graphs = sum(1 for e in entries if e["geonames_eligible"] == "yes")

    A("## Node classification (per glang's topology-zoo shape rules)\n")
    A(f"- **{total_real:,}** real internal nodes (become CORE routers).")
    A(f"- **{total_rj45:,}** Rj45 nodes (links to other networks; "
      f"`Internal == 0`).")
    A(f"- **{total_switch:,}** switch nodes (shared-medium fabrics; "
      f"`hyperedge == 1`).")
    A(f"- **{total_external:,}** edges connect to a network beyond the "
      f"graph (at least one Rj45 endpoint).")
    A("")

    A("## Real-node lat/lon coverage\n")
    A(f"- **{real_unloc_graphs}** of **{total_graphs}** graphs have at "
      f"least one real internal node missing lat/lon "
      f"(the Rj45/switch holes are excluded).")
    A(f"- Total real internal nodes lacking lat/lon: **{total_real_unloc:,}** "
      f"({total_real_unloc * 100.0 / max(total_real, 1):.1f}% of real nodes).")
    A("")

    A("## Geonames eligibility (upper bound for 100% population coverage)\n")
    A("Geonames offers two lookup paths.  A real internal node is "
      "*reachable* if at least one applies:\n")
    A("- **Reverse** (lat/lon → nearest populated place).  No country "
      "scope needed; coordinates disambiguate globally.")
    A("- **Forward** (place name → matching geonames record).  Needs a "
      "country scope, either the node's own `Country` attribute or a "
      "graph-level single-country context "
      "(`GeoExtent == \"Country\"` with a non-empty `GeoLocation`).\n")
    A("A graph is *geonames-eligible* when **every** real internal node is "
      "reachable by at least one path.  This is the upper bound on graphs "
      "that could plausibly have population data added to every internal "
      "node via the geonames API.\n")
    total_unreachable = sum(e["real_geonames_unreachable"] for e in entries)
    A(f"- **{geon_eligible_graphs}** of **{total_graphs}** graphs are "
      f"geonames-eligible "
      f"({geon_eligible_graphs * 100.0 / max(total_graphs, 1):.1f}%).")
    A(f"- Across the whole archive, **{total_unreachable:,}** real internal "
      f"nodes are unreachable by either path "
      f"({total_unreachable * 100.0 / max(total_real, 1):.2f}% of real nodes).")
    A("")

    # Measured (vs upper bound): the audit ran live geonames lookups
    # against the cities500 dataset for every real internal node.  A
    # node is *resolved* when it matched a geonames record (reverse by
    # lat/lon, or forward by name+country).  Forward matches can miss
    # even when the node was structurally reachable -- the label may
    # not appear in cities500 (small towns, abbreviations, alternate
    # spellings).  This section reports what actually matched.
    total_resolved   = sum(e["real_nodes_resolved"]   for e in entries)
    total_population = sum(e["real_total_population"] for e in entries)
    fully_resolved_graphs = sum(
        1 for e in entries
        if e["real_nodes"] > 0
        and e["real_nodes_resolved"] == e["real_nodes"])
    A("## Geonames coverage (measured against cities500)\n")
    A("Live cities500 lookups were run for every real internal node.  "
      "A node is *resolved* when geonames returned a record (reverse "
      "from lat/lon, or forward from name+country).  Resolved nodes "
      "carry a population figure on the `Real-node Populations` sheet.\n")
    A(f"- **{total_resolved:,}** of **{total_real:,}** real internal "
      f"nodes resolved "
      f"({total_resolved * 100.0 / max(total_real, 1):.1f}%).")
    A(f"- **{fully_resolved_graphs}** of **{total_graphs}** graphs have "
      f"100% of their real internal nodes resolved "
      f"({fully_resolved_graphs * 100.0 / max(total_graphs, 1):.1f}%).")
    A(f"- Aggregate population across resolved nodes: "
      f"**{total_population:,}**.")
    A("")

    # Partial-resolution to-do list: per-graph subsections listing the
    # unresolved real internal nodes in each partially-resolved graph,
    # so a human can work through them.  Sorted by completion %
    # descending (easiest finishes first).  Mirrors the `Partial
    # Resolution` workbook sheet.
    partials = []
    for e in entries:
        n = e.get("real_nodes", 0)
        k = e.get("real_nodes_resolved", 0)
        if isinstance(n, int) and isinstance(k, int) and 0 < k < n:
            partials.append((e["graph"], n, k))
    partials.sort(key=lambda t: (-t[2] / t[1], t[0]))
    if partials:
        A("## Partial-resolution to-do list\n")
        A(f"**{len(partials)}** graphs resolved most of their real "
          f"internal nodes but missed at least one.  Each subsection "
          f"below lists the unresolved labels alongside the lookup "
          f"method that failed (`forward-miss` -- a name+country lookup "
          f"that didn't match cities500; `unreachable` -- no lat/lon and "
          f"no usable name+country).  Listed in completion-percentage "
          f"order so the easiest graphs to finish appear first.\n")
        for g, n, k in partials:
            pct = k * 100.0 / n
            A(f"### {g} -- {k}/{n} resolved ({pct:.1f}%)\n")
            for row in all_enrichment.get(g, []):
                (_, nid, label, _slat, _slon,
                 _mc, matched_name, _pop, method) = row
                if matched_name:
                    continue
                disp = label if label else "(empty label)"
                A(f"- `{nid}` `{disp}` ({method})")
            A("")

    A("## Graph structure\n")
    A(f"- Multigraphs (at least one pair of parallel edges): **{len(multigraphs)}**.")
    A(f"- Total parallel edges across the archive: **{total_para:,}**.")
    A(f"- Total self-loops: **{total_self}**.")
    A(f"- Graphs with >1 connected component: "
      f"**{sum(1 for e in entries if e['components'] > 1)}**.")
    A(f"- Isolated (edgeless) nodes overall: **{total_iso}**.")
    A(f"- Edges whose source or target was not declared as a node: **{total_dang}**.")
    A("")

    A("## Labels\n")
    A(f"- Nodes without any `label` attribute: **{total_empty}**.")
    A(f"- Nodes whose label is whitespace only: **{total_ws}**.")
    A(f"- Nodes whose label contains control characters: **{total_ctrl}**.")
    A(f"- Graphs containing duplicate labels: "
      f"**{sum(1 for e in entries if e['duplicate_labels'] > 0)}**.")
    A(f"- Total duplicate-label occurrences: **{total_dup}** "
      "(counted as 1 per label that has ≥2 carriers within its graph).")
    A("")

    A("## Issue tally by category\n")
    for cat, c in sorted(category_counts.items(), key=lambda kv: -kv[1]):
        A(f"- `{cat}`: {c}")
    A("")

    A("## Issue tally by severity\n")
    for sev, c in sorted(severity_counts.items(), key=lambda kv: -kv[1]):
        A(f"- `{sev}`: {c}")
    A("")

    A("## Largest graphs\n")
    A("| Graph | Nodes | Edges | Multi? | Missing latlon |")
    A("|---|---:|---:|:---:|---:|")
    for e in most_nodes:
        A(f"| {e['graph']} | {e['nodes']} | {e['edges']} | "
          f"{e['multigraph?']} | {e['missing_lat_or_lon']} |")
    A("")

    A("## Worst location coverage (absolute count)\n")
    A("| Graph | Nodes | Missing | % |")
    A("|---|---:|---:|---:|")
    for e in most_missing:
        pct = e['missing_lat_or_lon'] * 100.0 / max(e['nodes'], 1)
        A(f"| {e['graph']} | {e['nodes']} | {e['missing_lat_or_lon']} | {pct:.1f}% |")
    A("")

    if multigraphs:
        A("## Top multigraphs by parallel-edge count\n")
        A("| Graph | Edges | Parallel pairs |")
        A("|---|---:|---:|")
        for e in most_multi:
            A(f"| {e['graph']} | {e['edges']} | {e['parallel_edges']} |")
        A("")

    A("## How to use the workbook\n")
    A("The companion workbook `zoo-audit.xlsx` has:")
    A("1. **Summary** — one row per graph with the same columns as the totals "
      "at the bottom of this report. The last row is the cross-archive total.")
    A("2. **Missing Locations** — every single node with a missing or invalid "
      "lat/lon across the whole corpus.")
    A("3. **Issues** — every issue found, labelled by severity (error / warn / "
      "info) and category, with the node or edge id in the detail.")
    A("4. **Multigraphs** — the subset of graphs with parallel edges, plus their "
      "edge counts.")
    A("5. **one sheet per graph** — full attribute table for every node and "
      "every edge; rows with problems are highlighted.")
    A("")
    A("## Notes on terminology\n")
    A("- \"Missing lat/lon\" counts a node if *either* of its Latitude or Longitude "
      "attributes is absent, empty, one of the sentinels `None`/`NaN`/`null`/`-`, or "
      "fails numeric coercion.")
    A("- \"Multigraph\" here means *observed* parallel edges, not the GraphML "
      "`parse.multiplegraph` flag (which is rarely set in the Zoo files).")
    A("- \"Duplicate labels\" counts **distinct labels** carried by 2+ nodes; for a "
      "count of *affected nodes* multiply by at least 2.")

    return "\n".join(lines) + "\n"


# ---- PDF rendering --------------------------------------------------
#
# A small markdown-to-PDF using reportlab. Enough to render *this*
# report cleanly: title, two heading levels, bullet lists, GitHub-
# flavoured tables, and inline emphasis (**bold**, `code`). Not a
# general markdown renderer.


def _md_inline_to_rl(text: str) -> str:
    """Convert inline markdown to the limited HTML reportlab accepts."""
    # Escape <, >, & first so we don't touch them inside markdown emphasis.
    out = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    out = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", out)
    out = re.sub(r"`([^`]+)`", r'<font face="Courier">\1</font>', out)
    return out


def render_pdf(md_text: str, out_path: Path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, ListFlowable,
                                    ListItem)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                 alignment=TA_LEFT, spaceAfter=6)
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], spaceBefore=14,
                        spaceAfter=6, textColor=colors.HexColor("#1f2a4e"))
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], spaceBefore=10,
                        spaceAfter=4, textColor=colors.HexColor("#3c4b7a"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], spaceAfter=4,
                          leading=14)
    mono = ParagraphStyle("Mono", parent=body, fontName="Courier",
                          fontSize=9, textColor=colors.HexColor("#555555"))

    story = []
    lines = md_text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if not stripped:
            story.append(Spacer(1, 4))
            i += 1
            continue
        # Title (first # line) vs subsequent ## / ###
        if stripped.startswith("# "):
            story.append(Paragraph(_md_inline_to_rl(stripped[2:]), title_style))
            i += 1
            continue
        if stripped.startswith("## "):
            story.append(Paragraph(_md_inline_to_rl(stripped[3:]), h1))
            i += 1
            continue
        if stripped.startswith("### "):
            story.append(Paragraph(_md_inline_to_rl(stripped[4:]), h2))
            i += 1
            continue
        # Tables: consecutive lines starting with |
        if stripped.startswith("|"):
            rows = []
            while i < len(lines) and lines[i].lstrip().startswith("|"):
                row = [c.strip() for c in lines[i].strip().strip("|").split("|")]
                rows.append(row)
                i += 1
            # Drop the separator row (--- |---|)
            filtered = [r for r in rows if not all(re.match(r"^:?-+:?$", c) for c in r)]
            if filtered:
                data = [[Paragraph(_md_inline_to_rl(c), body) for c in r]
                        for r in filtered]
                tbl = Table(data, hAlign="LEFT", repeatRows=1)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e6e8f0")),
                    ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOX",        (0, 0), (-1, -1), 0.25, colors.grey),
                    ("INNERGRID",  (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("VALIGN",     (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
                    ("TOPPADDING",    (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]))
                story.append(tbl)
                story.append(Spacer(1, 6))
            continue
        # Bulleted list
        if stripped.startswith("- "):
            items = []
            while i < len(lines) and lines[i].lstrip().startswith("- "):
                items.append(ListItem(Paragraph(_md_inline_to_rl(lines[i].lstrip()[2:]), body),
                                      leftIndent=12, bulletColor=colors.HexColor("#3c4b7a")))
                i += 1
            story.append(ListFlowable(items, bulletType="bullet",
                                      leftIndent=18, bulletFontSize=8,
                                      bulletOffsetY=-1))
            story.append(Spacer(1, 4))
            continue
        # Italic-copyright line ("*text*")
        if stripped.startswith("*") and stripped.endswith("*") and not stripped.startswith("**"):
            story.append(Paragraph(f'<i>{_md_inline_to_rl(stripped.strip("*"))}</i>',
                                   mono))
            i += 1
            continue
        # Regular paragraph
        story.append(Paragraph(_md_inline_to_rl(stripped), body))
        i += 1

    doc = SimpleDocTemplate(str(out_path), pagesize=A4,
                            leftMargin=20 * mm, rightMargin=20 * mm,
                            topMargin=18 * mm, bottomMargin=18 * mm,
                            title="Topology Zoo data-quality audit",
                            author="Eric Parsonage")
    doc.build(story)


# ---- main ----------------------------------------------------------

def main():
    files = sorted(ZOO_DIR.glob("*.graphml"))
    if not files:
        print(f"No graphml files under {ZOO_DIR}", file=sys.stderr)
        sys.exit(1)

    print(f"Processing {len(files)} graphml files from {ZOO_DIR}")

    print("Loading geonames cities500 index ...", flush=True)
    geo = GeonamesIndex.load()
    print(f"  {len(geo.records):,} populated places indexed", flush=True)

    entries = []         # summary row per graph
    all_issues = {}      # graph -> list of issue dicts
    all_missing = {}     # graph -> list of (nid, label, which)
    all_enrichment = {}  # graph -> list of enrichment rows (real nodes only)
    wb = Workbook()
    wb.remove(wb.active)

    # Pre-create summary/issues/missing sheets at the end, so per-graph
    # sheets sort nicely in the middle.
    for path in files:
        try:
            parsed = parse_graphml(path)
        except ET.ParseError as ex:
            print(f"  !! parse error in {path.name}: {ex}")
            entries.append({"graph": path.stem, "nodes": 0,
                            "real_nodes": 0, "rj45_nodes": 0,
                            "switch_nodes": 0, "edges": 0,
                            "external_edges": 0,
                            "edgedefault": "?", "multigraph?": "?",
                            "parallel_edges": 0, "self_loops": 0,
                            "missing_lat_or_lon": 0,
                            "real_missing_lat_or_lon": 0,
                            "real_missing_label": 0,
                            "real_missing_country": 0,
                            "real_geonames_unreachable": 0,
                            "geonames_eligible": "no",
                            "real_nodes_resolved": 0,
                            "real_total_population": 0,
                            "bad_coords": 0,
                            "isolated_nodes": 0, "components": 0,
                            "duplicate_labels": 0, "whitespace_labels": 0,
                            "empty_labels": 0, "control_char_labels": 0,
                            "dangling_endpoints": 0, "issues": 1})
            continue
        summary, issues, missing, enrichment = analyse(parsed, geo=geo)
        entries.append(summary)
        all_issues[path.stem]     = issues
        all_missing[path.stem]    = missing
        all_enrichment[path.stem] = enrichment
        write_per_graph_sheet(wb, parsed)

    # Insert rollup sheets at position 0..3
    # (write_workbook builds those; we carry wb separately to keep the
    # per-graph sheets we already created.)
    roll = write_workbook(entries, all_missing, all_issues, all_enrichment,
                          None)
    # Insert roll sheets at the front of wb (in reverse so order is right)
    for title in ("Multigraphs", "Partial Resolution",
                  "Real-node Populations", "Issues",
                  "Missing Locations", "Summary"):
        if title in roll.sheetnames:
            src = roll[title]
            dst = wb.create_sheet(title=title, index=0)
            for row in src.iter_rows(values_only=False):
                for c in row:
                    nc = dst.cell(row=c.row, column=c.column, value=c.value)
                    nc.font = c.font.copy()
                    nc.fill = c.fill.copy()
            autosize(dst)

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"  wrote {XLSX}")

    md = write_report(entries, all_issues, all_enrichment, None)
    REPORT.write_text(md)
    print(f"  wrote {REPORT}")

    render_pdf(md, PDF)
    print(f"  wrote {PDF}")


if __name__ == "__main__":
    main()
